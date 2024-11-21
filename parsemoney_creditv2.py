import imaplib
import email
from email.policy import default
import re
import chardet
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# CREDENTIALS
GMAIL_USER = "andifallihmalela@gmail.com"
GMAIL_PASSWORD = "capsikvvakpaqujx"

file_path = "trial2.xlsx" 

def fetch_labeled_emails():
    # Connect to Gmail
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_PASSWORD)
    mail.select('"[Gmail]/All Mail"')  # Access all emails, including labeled ones

    # Search for emails with the label "SMBC-Credit"
    status, messages = mail.search(None, 'X-GM-LABELS "SMBC-Credit"')
    if status != "OK" or not messages[0]:
        print("No labeled emails found!")
        return mail, []

    email_ids = messages[0].split()
    emails = []

    for email_id in email_ids:
        # Fetch the email
        status, msg_data = mail.fetch(email_id, "(RFC822)")
        if status == "OK":
            msg = email.message_from_bytes(msg_data[0][1], policy=default)
            emails.append((email_id, msg))

    return mail, emails

def remove_label(mail, email_id):
    # Remove the "SMBC-Credit" label after processing
    mail.store(email_id, '-X-GM-LABELS', 'SMBC-Credit')

def extract_details(msg):
    # Check the email's declared encoding
    body = msg.get_body(preferencelist=("plain", "html"))
    body_bytes = body.get_content().encode() if isinstance(body.get_content(), str) else body.get_content()

    # Detect charset dynamically
    detected = chardet.detect(body_bytes)
    charset = detected.get('encoding', 'utf-8')
    print(f"Detected charset: {charset}")

    try:
        body_decoded = body_bytes.decode(charset)
    except UnicodeDecodeError:
        print(f"Failed to decode using {charset}.")
        body_decoded = body_bytes.decode("utf-8", errors="replace")

    print(f"Decoded content (first 200 chars): {body_decoded[:200]}")

    # Extract details using regex
    date_match = re.search(r"◇利用日：(\d{4}/\d{2}/\d{2})", body_decoded)
    merchant_match = re.search(r"◇利用先：(.+)", body_decoded)
    transaction_match = re.search(r"◇利用取引：(.+)", body_decoded)
    amount_match = re.search(r"◇利用金額：(\d+)", body_decoded)

    # Combine merchant and transaction type into one
    merchant_transaction = f"{merchant_match.group(1).strip()} ({transaction_match.group(1).strip()})" if merchant_match and transaction_match else None

    return {
        "date": date_match.group(1).strip() if date_match else None,
        "merchant_transaction": merchant_transaction,
        "amount": amount_match.group(1).strip() if amount_match else None,
    }

def find_next_empty_row(sheet):
    # Check for the next empty row in the last sheet based on specific columns
    for row in range(1, sheet.max_row + 1):
        if all(sheet.cell(row=row, column=col).value is None for col in [1, 2, 3]):  # Columns A, B, C
            return row
    return sheet.max_row + 1

def update_excel(details):
    wb = load_workbook(file_path)
    sheet = wb.worksheets[-1]  # Get the last sheet in the workbook

    # Find the next empty row based on relevant columns
    next_row = find_next_empty_row(sheet)

    # Write details into columns (A, B, C) with formatting
    sheet.cell(row=next_row, column=1, value=details["date"]).number_format = "YYYY/MM/DD"  # Set date format
    sheet.cell(row=next_row, column=2, value=details["merchant_transaction"])
    sheet.cell(row=next_row, column=3, value=float(details["amount"]) if details["amount"].isdigit() else details["amount"])  # Convert to number if possible

    # Apply formatting (e.g., borders)
    thin_border = Border(
        left=Side(style="thin"), 
        right=Side(style="thin"), 
        top=Side(style="thin"), 
        bottom=Side(style="thin")
    )

    for col in range(1, 4):  # Apply borders to columns A, B, C
        cell = sheet.cell(row=next_row, column=col)
        cell.border = thin_border

        # Apply alignment for specific columns
        if col != 2:  # Center-align only columns A (date) and C (amount), skip column B
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook
    wb.save(file_path)
    print(f"Updated Excel on sheet '{sheet.title}' with: {details}")

# Main workflow
if __name__ == "__main__":
    mail, emails = fetch_labeled_emails()
    for email_id, msg in emails:
        details = extract_details(msg)
        if details:
            update_excel(details)
            # Remove the label after successful processing
            remove_label(mail, email_id)

    # Logout from the mail server
    mail.logout()
