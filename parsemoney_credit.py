import imaplib
import email
from email.policy import default
import re
import chardet
from openpyxl import load_workbook

# CREDENTIALS
GMAIL_USER = "andifallihmalela@gmail.com"
GMAIL_PASSWORD = "capsikvvakpaqujx"

def fetch_email():
    # Connect to Gmail via IMAP
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_PASSWORD)
    mail.select("inbox")  # Select the inbox folder

    # Search for emails from a specific sender
    status, messages = mail.search(None, 'FROM "statement@vpass.ne.jp"')
    if status != "OK":
        print("No messages found!")
        return None

    # Fetch the latest email
    latest_email_id = messages[0].split()[-1]
    status, msg_data = mail.fetch(latest_email_id, "(RFC822)")
    if status != "OK":
        print("Failed to fetch email!")
        return None

    # Parse the email
    msg = email.message_from_bytes(msg_data[0][1], policy=default)
    return msg

def extract_details(msg):
    # Check the email's declared encoding
    body = msg.get_body(preferencelist=("plain", "html"))
    body_bytes = body.get_content().encode() if isinstance(body.get_content(), str) else body.get_content()

    # Detect charset dynamically

    detected = chardet.detect(body_bytes)
    charset = detected.get('encoding', 'utf-8')
    print(f"Detected charset: {charset}")

    try:
        body_decoded = body_bytes.decode(charset)
    except UnicodeDecodeError:
        print(f"Failed to decode using {charset}.")
        body_decoded = body_bytes.decode("utf-8", errors="replace")

    print(f"Decoded content (first 200 chars): {body_decoded[:200]}")

    # Extract details using regex
    date_match = re.search(r"◇利用日：(\d{4}/\d{2}/\d{2})", body_decoded)
    merchant_match = re.search(r"◇利用先：(.+)", body_decoded)
    transaction_match = re.search(r"◇利用取引：(.+)", body_decoded)
    amount_match = re.search(r"◇利用金額：(\d+)", body_decoded)

    # Combine merchant and transaction type into one
    merchant_transaction = f"{merchant_match.group(1).strip()} ({transaction_match.group(1).strip()})" if merchant_match and transaction_match else None

    return {
        "date": date_match.group(1).strip() if date_match else None,
        "merchant_transaction": merchant_transaction,
        "amount": amount_match.group(1).strip() if amount_match else None,
    }



def update_excel(details):
    file_path = "/Users/andrew/Zed/proyek/parsemoney/transactions.xlsx" 
    wb = load_workbook(file_path)
    sheet = wb.active  # Select the first sheet

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Write details into columns
    sheet.cell(row=next_row, column=1, value=details["date"])
    sheet.cell(row=next_row, column=2, value=details["merchant_transaction"])
    sheet.cell(row=next_row, column=3, value=details["amount"])

    # Save the file
    wb.save(file_path)
    print(f"Updated Excel with: {details}")

# Main workflow
if __name__ == "__main__":
    msg = fetch_email()
    if msg:
        details = extract_details(msg)
        if details:
            update_excel(details)